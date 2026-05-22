"""
Parser for MPGU exam/credit session schedules.

Format: PDF or Excel with a transposed table layout:
  - Rows = date+time slots (with vertical cell text)
  - Columns = groups (from header row)
  - Cells = subject + type + teacher + room (may overflow into sub-tables)

Also handles scanned PDFs via Tesseract fallback.
"""
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

log = logging.getLogger(__name__)

DAYS_RU = {
    "понедельник", "вторник", "среда", "четверг",
    "пятница", "суббота", "воскресенье",
}

RU_MONTHS = {
    "января": "01", "февраля": "02", "марта": "03",
    "апреля": "04", "мая": "05", "июня": "06",
    "июля": "07", "августа": "08", "сентября": "09",
    "октября": "10", "ноября": "11", "декабря": "12",
}


@dataclass
class ExamEntry:
    date: str                       # "YYYY-MM-DD"
    time_start: str                 # "HH:MM"
    time_end: str | None            # "HH:MM" or None
    subject: str
    type: str                       # "exam" | "credit" | "unknown"
    teacher: str | None
    room: str | None
    groups: list[str] = field(default_factory=list)


# ─── Vertical cell decoder ────────────────────────────────────────────────────

def _decode_vertical(cell: str) -> tuple[str, str]:
    """
    Decode a cell with vertical (rotated) text.
    pdfplumber returns each character on its own line.
    Returns (kind, value) where kind in {'day','date','time','raw'}.
    """
    parts = [c.strip() for c in cell.split("\n") if c.strip()]
    fwd = "".join(parts)
    rev = "".join(reversed(parts))

    # Day of week
    for t in (fwd, rev):
        if t.lower() in DAYS_RU:
            return "day", t.lower()

    # Date: DD.MM.YYYY or DD/MM/YYYY
    for t in (fwd, rev):
        m = re.search(r"(\d{2})[./](\d{2})[./](\d{4})", t)
        if m:
            return "date", f"{m.group(3)}-{m.group(2)}-{m.group(1)}"

    # Time: HH:MM–HH:MM or HH.MM-HH.MM
    for t in (fwd, rev):
        m = re.search(r"(\d{2})[.:](\d{2})\s*[-—]\s*(\d{2})[.:](\d{2})", t)
        if m:
            return "time", f"{m.group(1)}:{m.group(2)}-{m.group(3)}:{m.group(4)}"

    return "raw", fwd or rev


def _parse_date_text(text: str) -> str | None:
    """Parse a date from natural language text, e.g. '15 января 2026'."""
    m = re.search(r"(\d{1,2})\s+(" + "|".join(RU_MONTHS) + r")\s+(\d{4})", text, re.I)
    if m:
        d = m.group(1).zfill(2)
        mo = RU_MONTHS[m.group(2).lower()]
        y = m.group(3)
        return f"{y}-{mo}-{d}"
    m2 = re.search(r"(\d{2})[./](\d{2})[./](\d{4})", text)
    if m2:
        return f"{m2.group(3)}-{m2.group(2)}-{m2.group(1)}"
    return None


# ─── Cell content parser ─────────────────────────────────────────────────────

def _parse_cell_content(lines: list[str]) -> dict | None:
    """Extract subject, type, teacher, room from a list of text lines."""
    lines = [ln.strip() for ln in lines if ln.strip()]
    if not lines:
        return None

    subject_parts: list[str] = []
    type_: str = "unknown"
    teacher: str | None = None
    room: str | None = None

    for line in lines:
        up = line.upper()
        # Standalone type keyword lines (no other subject content)
        stripped = re.sub(r"[\s().]", "", up)
        if stripped in {"ЭКЗАМЕН", "ЭКЗ", "КОНС", "КОНСУЛЬТАЦИЯ"}:
            type_ = "exam"
        elif stripped in {"ЗАЧЁТ", "ЗАЧЕТ", "ЗАЧ"}:
            type_ = "credit"
        elif re.search(r"ауд\.?\s*[\d\w/\\-]+", line, re.I):
            m = re.search(r"ауд\.?\s*([\d\w/\\-]+)", line, re.I)
            if m:
                room = m.group(1)
        elif re.search(
            r"\b(проф|доц|ст\.?\s*пр|асс|преп)[.\s]"
            r"|\b(профессор|доцент|преподаватель|старший\s+преподаватель)\b",
            line, re.I,
        ):
            teacher = line
        elif re.search(r"[А-ЯЁ][а-яё]+\s+[А-ЯЁ]\.[А-ЯЁ]\.", line):
            teacher = line
        elif not re.search(r"^\(|ауд\.|зачет|экзамен", line, re.I):
            subject_parts.append(line)

    subject = " ".join(subject_parts).strip()
    if not subject:
        return None

    # Post-process: detect type keywords embedded in the subject line
    sub_up = subject.upper()
    if type_ == "unknown":
        if re.search(r"\bЭКЗ\b|ЭКЗАМЕН", sub_up):
            type_ = "exam"
        elif re.search(r"\bЗАЧ\b|ЗАЧЁТ|ЗАЧЕТ", sub_up):
            type_ = "credit"
        elif re.search(r"\bКОНС\b|КОНСУЛЬТАЦ", sub_up):
            type_ = "exam"

    return {
        "subject": subject,
        "type": type_,
        "teacher": teacher,
        "room": room,
    }


# ─── PDF parser ──────────────────────────────────────────────────────────────

def _parse_pdf(path: str) -> list[ExamEntry]:
    try:
        import pdfplumber
    except ImportError:
        log.warning("pdfplumber not available")
        return []

    entries: list[ExamEntry] = []
    # Thread group X-ranges across pages: some PDFs (e.g. philology) put the
    # group header only on page 2; subsequent data pages need the inherited ranges.
    inherited_x_ranges: list[tuple[float, float, str]] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            page_entries, page_x_ranges = _extract_page(page, inherited_x_ranges)
            if page_x_ranges:
                inherited_x_ranges = page_x_ranges
            entries.extend(page_entries)

    # If text extraction failed (scanned PDF), try Tesseract
    if not entries:
        entries = _parse_pdf_ocr(path)

    return entries


def _extract_page(
    page,
    inherited_group_x_ranges: list[tuple[float, float, str]] | None = None,
) -> tuple[list[ExamEntry], list[tuple[float, float, str]]]:
    """
    Returns (entries, group_x_ranges).
    group_x_ranges is non-empty when this page has a group header row; the
    caller threads it into subsequent pages that lack their own header.
    """
    ts = {"vertical_strategy": "lines", "horizontal_strategy": "lines"}
    try:
        finders = list(page.find_tables(ts))
    except Exception:
        return [], []
    if not finders:
        return [], []

    tables = [(f.bbox, f.extract()) for f in finders]

    # ── Step 1: find main schedule table (has group headers) ──
    main_idx: int | None = None
    header_row_idx: int = 0
    group_col_map: list[tuple[int, str]] = []  # [(col_idx, group_name)]

    for ti, (bbox, table) in enumerate(tables):
        for ri, row in enumerate(table):
            cells = [str(c or "").strip() for c in row]
            # Standard Cyrillic group codes like БОИ35-ИПЛ2402
            grps = [
                (ci, c) for ci, c in enumerate(cells)
                if re.search(r"[А-ЯЁ]{2,}\d{2}", c) and len(c) < 60
            ]
            if not grps:
                # Numeric group codes (philology-style: "101", "102" …).
                # Require ≥3 matches per row to avoid false positives from
                # dates, room numbers, etc.
                numeric_grps = [
                    (ci, c) for ci, c in enumerate(cells)
                    if re.match(r"^\d{3,4}$", c)
                ]
                if len(numeric_grps) >= 3:
                    grps = numeric_grps
            if grps:
                main_idx = ti
                header_row_idx = ri
                group_col_map = grps
                break
        if main_idx is not None:
            break

    if main_idx is None or not group_col_map:
        # No group header on this page — fall back to inherited X ranges if
        # available (e.g. philology data pages 3-10 after the header on page 2)
        if inherited_group_x_ranges:
            return _extract_with_inherited_x(page, tables, inherited_group_x_ranges), []
        return [], []

    main_bbox, main_table = tables[main_idx]

    # ── Step 2: extract date/time from main table rows ──
    # row_meta[ri] = {'date': ..., 'time_start': ..., 'time_end': ...}
    row_meta: dict[int, dict] = {}
    cur_date: str | None = None
    cur_time: str | None = None

    for ri, row in enumerate(main_table):
        if ri <= header_row_idx:
            continue
        for cell in row:
            if not cell:
                continue
            s = str(cell).strip()
            if "\n" in s and len(s) < 100:
                kind, val = _decode_vertical(s)
                if kind == "date":
                    cur_date = val
                elif kind == "time":
                    cur_time = val
        if cur_date or cur_time:
            row_meta[ri] = {
                "date": cur_date,
                "time": cur_time,
            }

    # ── Step 3: extract exam content from main table cells ──
    entries_from_main = _entries_from_main_table(
        main_table, header_row_idx, group_col_map, row_meta
    )

    # ── Step 4: build X-ranges once, share with sub-table extractor ──
    group_x_ranges = _get_group_x_ranges(page, group_col_map, main_bbox)

    # ── Step 5: extract exam content from overflow sub-tables ──
    # Sub-tables are positioned within the grid; match by Y-overlap to row_meta
    entries_from_sub = _entries_from_sub_tables(
        tables, main_idx, main_bbox, main_table, header_row_idx,
        group_col_map, row_meta, page, group_x_ranges,
    )

    # Merge, prefer sub-table entries (they have full content)
    entries = entries_from_sub if entries_from_sub else entries_from_main
    return entries, group_x_ranges


def _entries_from_main_table(
    table: list[list],
    header_row_idx: int,
    group_col_map: list[tuple[int, str]],
    row_meta: dict[int, dict],
) -> list[ExamEntry]:
    entries = []
    # Carry forward date across rows
    cur_date = None

    for ri, row in enumerate(table):
        if ri <= header_row_idx:
            continue
        meta = row_meta.get(ri, {})
        if meta.get("date"):
            cur_date = meta["date"]
        date = cur_date
        time_str = meta.get("time")
        time_start, time_end = _split_time(time_str)

        if not date:
            continue

        cells = [str(c or "").strip() for c in row]
        for col_idx, group_name in group_col_map:
            if col_idx >= len(cells):
                continue
            cell = cells[col_idx]
            if not cell or len(cell) < 3:
                continue
            content = _parse_cell_content(cell.split("\n"))
            if content:
                entries.append(ExamEntry(
                    date=date,
                    time_start=time_start or "",
                    time_end=time_end,
                    subject=content["subject"],
                    type=content["type"],
                    teacher=content["teacher"],
                    room=content["room"],
                    groups=[_clean_group_name(group_name)],
                ))
    return entries


def _entries_from_sub_tables(
    tables: list[tuple],
    main_idx: int,
    main_bbox: tuple,
    main_table: list[list],
    header_row_idx: int,
    group_col_map: list[tuple[int, str]],
    row_meta: dict[int, dict],
    page,
    group_x_ranges: list[tuple[float, float, str]] | None = None,
) -> list[ExamEntry]:
    """
    Match sub-tables to (group, date, time) by spatial position.
    """
    # Accept pre-computed ranges from _extract_page to avoid duplicate work
    if group_x_ranges is None:
        group_x_ranges = _get_group_x_ranges(page, group_col_map, main_bbox)
    if not group_x_ranges:
        return []

    # Get row y-ranges from main table structure
    row_y_ranges = _get_row_y_ranges(page, main_table, header_row_idx, main_bbox, row_meta)

    entries: list[ExamEntry] = []
    cur_date = None

    for ri_ordered, (y0, y1, date, time_str) in enumerate(row_y_ranges):
        if date:
            cur_date = date
        effective_date = cur_date

        for sub_idx, (sub_bbox, sub_table) in enumerate(tables):
            if sub_idx == main_idx:
                continue
            sx0, sy0, sx1, sy1 = sub_bbox
            # Y-overlap check
            if sy1 < y0 or sy0 > y1:
                continue
            # Find matching group by X-overlap
            sub_cx = (sx0 + sx1) / 2
            group_name = _find_group_for_x(sub_cx, group_x_ranges)
            if not group_name:
                continue

            # Extract content from sub-table
            lines = []
            for row in sub_table:
                for cell in row:
                    if cell:
                        lines.extend(str(cell).split("\n"))

            content = _parse_cell_content(lines)
            if not content or not effective_date:
                continue

            time_start, time_end = _split_time(time_str)
            entries.append(ExamEntry(
                date=effective_date,
                time_start=time_start or "",
                time_end=time_end,
                subject=content["subject"],
                type=content["type"],
                teacher=content["teacher"],
                room=content["room"],
                groups=[_clean_group_name(group_name)],
            ))

    return entries


def _get_group_x_ranges(page, group_col_map, main_bbox) -> list[tuple[float, float, str]]:
    """
    Get (x_left, x_right, group_name) for each group column.

    Strategy: find group-code words in the header area, sort both group_col_map
    (by col_idx) and found words (by x) and zip them in order.  This handles
    duplicate codes like two "ВОИ34-ИОВ2503" groups (п/г 1 and п/г 2).
    """
    words = page.extract_words()
    main_x0, main_y0, main_x1, main_y1 = main_bbox
    header_y_max = main_y0 + 100

    # Choose regex based on whether the map uses numeric codes (philology-style)
    is_numeric = all(re.match(r"^\d{3,4}$", name) for _, name in group_col_map)
    if is_numeric:
        group_code_re = re.compile(r"^\d{3,4}$")
        header_group_words = sorted(
            [w for w in words if group_code_re.match(w["text"]) and w["top"] <= header_y_max],
            key=lambda w: w["x0"],
        )
    else:
        group_code_re = re.compile(r"[А-ЯЁA-Z]{2,}\d{2}")
        header_group_words = sorted(
            [w for w in words if group_code_re.search(w["text"]) and w["top"] <= header_y_max],
            key=lambda w: w["x0"],
        )

    # Sort groups by column index (left-to-right order)
    sorted_groups = sorted(group_col_map, key=lambda x: x[0])

    if not header_group_words:
        return []

    # Match in order: i-th group (sorted by col) → i-th header word (sorted by x)
    result = []
    for i, (col_idx, group_name) in enumerate(sorted_groups):
        if i >= len(header_group_words):
            break
        w = header_group_words[i]
        x_center = (w["x0"] + w["x1"]) / 2
        # Estimate column width as distance to next group word / 2
        if i + 1 < len(header_group_words):
            next_cx = (header_group_words[i + 1]["x0"] + header_group_words[i + 1]["x1"]) / 2
            half_width = (next_cx - x_center) / 2
        else:
            half_width = (main_x1 - x_center) / 2
        result.append((x_center - half_width, x_center + half_width, group_name))

    return result


def _get_row_y_ranges(
    page, main_table, header_row_idx, main_bbox, row_meta
) -> list[tuple[float, float, str | None, str | None]]:
    """
    Approximate y-ranges for data rows using equal distribution within table height.
    Returns [(y0, y1, date, time)] for each data row.
    """
    x0, y0, x1, y1 = main_bbox
    data_rows = [ri for ri in range(len(main_table)) if ri > header_row_idx]
    if not data_rows:
        return []

    row_height = (y1 - y0) / len(main_table)
    ranges = []
    for ri in data_rows:
        ry0 = y0 + ri * row_height
        ry1 = ry0 + row_height
        meta = row_meta.get(ri, {})
        ranges.append((ry0, ry1, meta.get("date"), meta.get("time")))
    return ranges


def _extract_with_inherited_x(
    page,
    tables: list[tuple],
    group_x_ranges: list[tuple[float, float, str]],
) -> list[ExamEntry]:
    """
    Extract entries from pages that lack a group header row.

    Used for philology-style PDFs where the group header (numeric codes 101-111)
    only appears on page 2, while pages 3-10 are pure data pages with the same
    physical column layout.  We estimate each column's X-center from the table
    bounding box and map it to the nearest inherited group X-range.
    """
    entries: list[ExamEntry] = []
    for bbox, table in tables:
        if not table:
            continue
        tbl_x0, tbl_y0, tbl_x1, tbl_y1 = bbox
        n_cols = max((len(r) for r in table), default=0)
        if n_cols < 3:
            continue

        col_width = (tbl_x1 - tbl_x0) / n_cols
        cur_date: str | None = None
        cur_time: str | None = None

        for row in table:
            cells = [str(c or "").strip() for c in row]

            # First two columns carry date/time (vertical or plain text)
            for ci in range(min(2, len(cells))):
                cell = cells[ci]
                if not cell:
                    continue
                if "\n" in cell and len(cell) < 100:
                    kind, val = _decode_vertical(cell)
                    if kind == "date":
                        cur_date = val
                    elif kind == "time":
                        cur_time = val
                else:
                    date = _parse_date_text(cell)
                    if date:
                        cur_date = date
                    m = re.search(r"(\d{2})[.:](\d{2})\s*[-—]\s*(\d{2})[.:](\d{2})", cell)
                    if m:
                        cur_time = f"{m.group(1)}:{m.group(2)}-{m.group(3)}:{m.group(4)}"

            if not cur_date:
                continue

            time_start, time_end = _split_time(cur_time)

            for ci in range(2, len(cells)):
                cell = cells[ci]
                if not cell or len(cell) < 3:
                    continue
                col_cx = tbl_x0 + (ci + 0.5) * col_width
                group_name = _find_group_for_x(col_cx, group_x_ranges)
                if not group_name:
                    continue
                content = _parse_cell_content(cell.split("\n"))
                if not content:
                    continue
                entries.append(ExamEntry(
                    date=cur_date,
                    time_start=time_start or "",
                    time_end=time_end,
                    subject=content["subject"],
                    type=content["type"],
                    teacher=content["teacher"],
                    room=content["room"],
                    groups=[_clean_group_name(group_name)],
                ))

    return entries


def _find_group_for_x(cx: float, group_x_ranges) -> str | None:
    best_overlap = 0
    best_group = None
    for x0, x1, group_name in group_x_ranges:
        overlap = min(cx, x1) - max(cx, x0)
        if cx >= x0 and cx <= x1:
            return group_name
        dist = abs(cx - (x0 + x1) / 2)
        if best_group is None or dist < best_overlap:
            best_overlap = dist
            best_group = group_name
    return best_group


def _split_time(time_str: str | None) -> tuple[str | None, str | None]:
    if not time_str:
        return None, None
    m = re.match(r"(\d{2}:\d{2})-(\d{2}:\d{2})", time_str)
    if m:
        return m.group(1), m.group(2)
    return time_str, None


def _clean_group_name(name: str) -> str:
    # Strip newlines and trailing group labels like "\nГр. 101"
    name = re.sub(r"\s*\n.*", "", name)
    # Remove room number suffix like "(101)"
    return re.sub(r"\s*\(\d{3}\)\s*$", "", name).strip()


# ─── OCR fallback for scanned PDFs ───────────────────────────────────────────

_OCR_MAX_PAGES = 2      # skip OCR for PDFs longer than this (scanned single-page exams OK)
_OCR_PAGE_TIMEOUT = 30  # seconds per page before giving up

# Latin characters that Tesseract confuses with Cyrillic lookalikes.
# Only uppercase matters because we normalize to uppercase before matching.
_LATIN_TO_CYR = str.maketrans("HCOAEPBXKM", "НСОАЕРВХКМ")

# Group code pattern that tolerates OCR artifact З (Cyrillic Ze) instead of
# digit 3: e.g. "ВОКЗ4-МДЭ2501" is "ВОК34-МДЭ2501" after cleanup.
_OCR_GROUP_RE = re.compile(r"[А-ЯЁA-Z]{2,4}[\dЗ]{2}-[А-ЯЁA-Z]{2,4}\d{4}")

# Exam trigger: word-bounded so «ЭКЗАМЕНАЦИОННОЙ» does not fire.
# Applied to _norm_ocr() output, so Latin lookalikes (KOHC → КОНС) are
# already replaced before this regex runs.
_OCR_EXAM_KW = re.compile(
    r"\bЗАЧЁТ\b|\bЗАЧЕТ\b|\bЭКЗАМЕН\b|\bЗАЧ\b|\bЭКЗ\b|\bКОНС\b"
)


def _norm_ocr(text: str) -> str:
    """Upper-case and replace Latin lookalikes with Cyrillic for trigger matching."""
    return text.upper().translate(_LATIN_TO_CYR)


_OCR_TABLE_HEADER_RE = re.compile(
    r"дата\s*\|?\s*время|день\s+недели|№\s*п/п|\|\s*время\s*\|",
    re.IGNORECASE,
)


def _clean_ocr_line(line: str) -> str:
    """
    Strip leading OCR garbage from a line.
    Rotated-cell text (dates, row labels) often bleeds into the adjacent content
    cell as a short garbled prefix: 'ЕЯ ТЕОРИЯ…', '58 ЦИФРОВЫЕ…', 'x 28 ТЕОРИЯ…'.
    Pattern: one or more groups of (short uppercase blob OR non-alpha noise)
    followed by whitespace, anchored at the start.

    Returns empty string for table-header fragments ('дата | Время …').
    """
    if _OCR_TABLE_HEADER_RE.search(line):
        return ""
    return re.sub(r"^(?:(?:[\W\d]+|[А-ЯЁA-Z]{1,2})\s+)+", "", line)


def _parse_pdf_ocr(path: str) -> list[ExamEntry]:
    """Fallback: Tesseract OCR for image-based exam schedule PDFs."""
    try:
        from pdf2image import convert_from_path, pdfinfo_from_path
        import pytesseract
    except ImportError:
        return []

    try:
        info = pdfinfo_from_path(path)
        total_pages = info.get("Pages", 1)
    except Exception:
        total_pages = 1

    if total_pages > _OCR_MAX_PAGES:
        log.info("OCR skipped: %d-page PDF too large for OCR (%s)", total_pages, path)
        return []

    log.info("Falling back to OCR for exam schedule (%d pages): %s", total_pages, path)
    entries: list[ExamEntry] = []
    images = convert_from_path(path, dpi=150)  # 150 dpi: 4× faster than 200

    # Thread group code across pages: the cover page often has the group in the
    # header while subsequent pages carry the actual exam rows.
    ocr_group: str | None = None

    for img in images:
        try:
            text = pytesseract.image_to_string(
                img,
                lang="rus+eng",
                config="--psm 6 --oem 1",  # oem 1 = LSTM only (faster)
                timeout=_OCR_PAGE_TIMEOUT,
            )
        except RuntimeError:
            log.warning("OCR timeout on page in %s, skipping", path)
            continue
        page_entries = _parse_ocr_text(text, ocr_group=ocr_group)
        # Carry forward the first group code found on any page
        if not ocr_group and page_entries and page_entries[0].groups:
            ocr_group = page_entries[0].groups[0]
        # Back-fill earlier entries on the same page that got no group
        for e in page_entries:
            if not e.groups and ocr_group:
                e.groups = [ocr_group]
        entries.extend(page_entries)

    return entries


def _parse_ocr_text(text: str, ocr_group: str | None = None) -> list[ExamEntry]:
    """
    Parse exam entries from raw OCR text.

    ocr_group: pre-extracted group code (passed from _parse_pdf_ocr so it
    is available across all pages of the same document).
    """
    entries: list[ExamEntry] = []
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    # Try to extract group code from this page's text if not already known.
    # _norm_ocr replaces Latin lookalikes, then _OCR_GROUP_RE finds the code.
    # Digit 3 is sometimes OCR'd as Cyrillic З — strip it back after match.
    if not ocr_group:
        m = _OCR_GROUP_RE.search(_norm_ocr(text))
        if m:
            parts = m.group(0).upper().split("-", 1)
            parts[0] = parts[0].replace("З", "3")
            ocr_group = "-".join(parts)

    cur_date: str | None = None

    i = 0
    while i < len(lines):
        line = lines[i]
        date = _parse_date_text(line)
        if date:
            cur_date = date

        # Match both full-word keywords and inline abbreviations (ЭКЗ., КОНС., ЗАЧ.).
        # _norm_ocr handles mixed-script OCR artifacts: KOHC → КОНС, etc.
        if cur_date and _OCR_EXAM_KW.search(_norm_ocr(line)):
            # 1 line back for multi-line subject; 3 lines forward for time/teacher/room.
            # Smaller window prevents bleeding into the next exam entry.
            raw_window = lines[max(0, i - 1): i + 4]
            # Strip leading OCR garbage (garbled rotated-cell text) from each line,
            # and drop lines with fewer than 3 alpha chars (pure noise like "5a", "= 12").
            entry_lines = [
                _clean_ocr_line(l)
                for l in raw_window
                if len(re.findall(r"[А-ЯЁа-яёA-Za-z]", l)) >= 3
            ]
            if not entry_lines:
                i += 1
                continue
            content = _parse_cell_content(entry_lines)
            if content:
                # Re-check type on normalised subject to handle OCR Latin/Cyrillic
                # confusion (e.g. KOHC not caught by _parse_cell_content's regex).
                if content["type"] == "unknown":
                    ns = _norm_ocr(content["subject"])
                    if re.search(r"\bЭКЗАМЕН\b|\bЭКЗ\b", ns):
                        content["type"] = "exam"
                    elif re.search(r"\bЗАЧЁТ\b|\bЗАЧЕТ\b|\bЗАЧ\b", ns):
                        content["type"] = "credit"
                    elif re.search(r"\bКОНС\b|\bКОНСУЛЬТАЦ\b", ns):
                        content["type"] = "exam"
                # Search time in the raw (unfiltered) window so we don't miss
                # lines like "= 12.00 a" that were filtered out by alpha count.
                time_match = re.search(r"(\d{2}[.:]\d{2})", " ".join(raw_window))
                time_start = time_match.group(1).replace(".", ":") if time_match else ""
                entries.append(ExamEntry(
                    date=cur_date,
                    time_start=time_start,
                    time_end=None,
                    subject=content["subject"],
                    type=content["type"],
                    teacher=content["teacher"],
                    room=content["room"],
                    groups=[ocr_group] if ocr_group else [],
                ))
        i += 1

    return entries


# ─── Excel parser ─────────────────────────────────────────────────────────────

def _parse_excel(path: str) -> list[ExamEntry]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    wb = load_workbook(path, data_only=True)
    entries: list[ExamEntry] = []
    for ws in wb.worksheets:
        entries.extend(_parse_excel_sheet(ws))
    return entries


def _parse_excel_sheet(ws) -> list[ExamEntry]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row with group names
    header_row_idx = None
    group_col_map: list[tuple[int, str]] = []

    for ri, row in enumerate(rows):
        cells = [str(c or "").strip() for c in row]
        grps = [
            (ci, c) for ci, c in enumerate(cells)
            if re.search(r"[А-ЯЁ]{2,}\d{2}", c) and len(c) < 60
        ]
        if grps:
            header_row_idx = ri
            group_col_map = grps
            break

    if header_row_idx is None:
        return []

    entries = []
    cur_date = None
    cur_time = None

    for row in rows[header_row_idx + 1:]:
        cells = [str(c or "").strip() for c in row]
        # Scan for date/time in non-group columns
        for ci, cell in enumerate(cells):
            if not cell:
                continue
            date = _parse_date_text(cell)
            if date:
                cur_date = date
            m = re.search(r"(\d{2})[.:](\d{2})\s*[-—]\s*(\d{2})[.:](\d{2})", cell)
            if m:
                cur_time = f"{m.group(1)}:{m.group(2)}-{m.group(3)}:{m.group(4)}"

        if not cur_date:
            continue

        time_start, time_end = _split_time(cur_time)
        for col_idx, group_name in group_col_map:
            if col_idx >= len(cells):
                continue
            cell = cells[col_idx]
            content = _parse_cell_content(cell.split("\n") if "\n" in cell else [cell])
            if content:
                entries.append(ExamEntry(
                    date=cur_date,
                    time_start=time_start or "",
                    time_end=time_end,
                    subject=content["subject"],
                    type=content["type"],
                    teacher=content["teacher"],
                    room=content["room"],
                    groups=[_clean_group_name(group_name)],
                ))
    return entries


# ─── DOCX parser ─────────────────────────────────────────────────────────────

def _parse_docx(path: str) -> list[ExamEntry]:
    try:
        from docx import Document
    except ImportError:
        log.warning("python-docx not available")
        return []

    try:
        doc = Document(path)
    except Exception as e:
        log.warning("Cannot open DOCX %s: %s", path, e)
        return []

    entries: list[ExamEntry] = []
    for table in doc.tables:
        rows = _extract_docx_rows(table)
        entries.extend(_parse_docx_exam_table(rows))
    return entries


def _extract_docx_rows(table) -> list[list[str]]:
    """Extract rows, collapsing merged cells to their text value."""
    rows = []
    for row in table.rows:
        rows.append([cell.text.strip() for cell in row.cells])
    return rows


def _parse_docx_exam_table(rows: list[list[str]]) -> list[ExamEntry]:
    if not rows:
        return []

    # Find header row with group name codes like "АА34-ГРП2501"
    header_row_idx: int | None = None
    group_col_map: list[tuple[int, str]] = []

    for ri, row in enumerate(rows[:15]):
        grps = [
            (ci, c) for ci, c in enumerate(row)
            if re.search(r"[А-ЯЁ]{2,}\d{2}", c) and len(c) < 80
        ]
        if grps:
            header_row_idx = ri
            group_col_map = _dedup_group_cols(grps)
            break

    if header_row_idx is None:
        return []

    entries: list[ExamEntry] = []
    cur_date: str | None = None
    cur_time: str | None = None
    # Track per-column previous text to detect merged cells
    prev_row: list[str] = []

    for row in rows[header_row_idx + 1:]:
        # Detect new date/time in non-group columns
        group_cols = {ci for ci, _ in group_col_map}
        for ci, cell in enumerate(row):
            if not cell or ci in group_cols:
                continue
            # Skip if this is a merged cell repeat
            if prev_row and ci < len(prev_row) and cell == prev_row[ci]:
                continue
            date = _parse_date_text(cell)
            if date:
                cur_date = date
            m = re.search(r"(\d{1,2})[.:](\d{2})\s*[-—]\s*(\d{1,2})[.:](\d{2})", cell)
            if m:
                cur_time = (
                    f"{int(m.group(1)):02d}:{m.group(2)}"
                    f"-{int(m.group(3)):02d}:{m.group(4)}"
                )

        if not cur_date:
            prev_row = row
            continue

        time_start, time_end = _split_time(cur_time)

        for col_idx, group_name in group_col_map:
            if col_idx >= len(row):
                continue
            cell = row[col_idx]
            if not cell:
                continue
            # Skip merged-cell repeats in group columns too
            if prev_row and col_idx < len(prev_row) and cell == prev_row[col_idx]:
                continue
            lines = [ln.strip() for ln in cell.split("\n") if ln.strip()]
            content = _parse_cell_content(lines)
            if content:
                entries.append(ExamEntry(
                    date=cur_date,
                    time_start=time_start or "",
                    time_end=time_end,
                    subject=content["subject"],
                    type=content["type"],
                    teacher=content["teacher"],
                    room=content["room"],
                    groups=[_clean_group_name(group_name)],
                ))

        prev_row = row

    return entries


def _dedup_group_cols(
    group_col_map: list[tuple[int, str]],
) -> list[tuple[int, str]]:
    """Remove duplicate group names (merged header cells) keeping first occurrence."""
    seen: set[str] = set()
    result = []
    for col_idx, name in group_col_map:
        clean = _clean_group_name(name)
        if clean not in seen:
            seen.add(clean)
            result.append((col_idx, name))
    return result


# ─── Format detection ────────────────────────────────────────────────────────

def _detect_bytes_format(data: bytes, hint_ext: str = "") -> str:
    """Detect file format from magic bytes; hint_ext overrides if reliable."""
    clean = hint_ext.lstrip(".").lower()
    if clean in {"pdf", "xlsx", "xls", "docx", "doc"}:
        return clean

    sig = data[:8]
    if sig[:4] == b"%PDF":
        return "pdf"
    if sig == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"
    if sig[:4] == b"PK\x03\x04":
        import zipfile, io
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                names = zf.namelist()
            if "[Content_Types].xml" in names:
                if any(n.startswith("xl/") for n in names):
                    return "xlsx"
                if any(n.startswith("word/") for n in names):
                    return "docx"
                return "xlsx"
            return "zip"
        except Exception:
            return "zip"
    if data[:5] in (b"<!DOC", b"<html") or data[:1] == b"<":
        return "html"
    return "unknown"


def _parse_zip_bytes(data: bytes) -> list[ExamEntry]:
    """Extract and parse each supported file inside a ZIP archive."""
    import zipfile, io
    entries: list[ExamEntry] = []
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for name in zf.namelist():
                if name.endswith("/"):
                    continue
                ext = Path(name).suffix.lower()
                if ext not in {".pdf", ".xlsx", ".xls", ".docx", ".doc"}:
                    continue
                try:
                    file_data = zf.read(name)
                    found = parse_exam_bytes(file_data, hint_ext=ext)
                    log.info("  ZIP entry %s → %d entries", name, len(found))
                    entries.extend(found)
                except Exception as e:
                    log.warning("  ZIP entry %s: %s", name, e)
    except Exception as e:
        log.warning("ZIP open error: %s", e)
    return entries


# ─── Public API ──────────────────────────────────────────────────────────────

def parse_exam_bytes(data: bytes, hint_ext: str = "") -> list[ExamEntry]:
    """Parse exam entries from raw bytes, auto-detecting format."""
    fmt = _detect_bytes_format(data, hint_ext)
    if fmt in {"html", "unknown"}:
        log.debug("Unsupported or HTML content (fmt=%s), skipping", fmt)
        return []
    if fmt == "zip":
        return _parse_zip_bytes(data)

    import tempfile, os
    suffix = "." + fmt
    fd, tmp_path = tempfile.mkstemp(suffix=suffix)
    try:
        os.write(fd, data)
        os.close(fd)
        return parse_exam_file(tmp_path)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def parse_exam_file(path: str) -> list[ExamEntry]:
    """Parse exam/credit schedule from a PDF, Excel, or DOCX file."""
    ext = Path(path).suffix.lower()
    if ext in {".xlsx", ".xls"}:
        return _parse_excel(path)
    elif ext in {".docx", ".doc"}:
        return _parse_docx(path)
    else:
        return _parse_pdf(path)


def entries_to_dicts(entries: list[ExamEntry]) -> list[dict]:
    return [
        {
            "date": e.date,
            "time_start": e.time_start,
            "time_end": e.time_end,
            "subject": e.subject,
            "type": e.type,
            "teacher": e.teacher,
            "room": e.room,
            "groups": e.groups,
        }
        for e in sorted(entries, key=lambda x: (x.date, x.time_start))
    ]
