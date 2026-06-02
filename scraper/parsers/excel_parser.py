"""Парсер Excel (.xlsx/.xls) расписаний МПГУ."""
import re

import openpyxl

from scraper.parsers.base import BaseParser, ParseResult
from scraper.normalizer.schedule_normalizer import (
    normalize_day, normalize_lesson_type, normalize_time,
    make_schedule_skeleton, lesson_obj, extract_subgroup, date_str_to_weekday,
)


class ExcelParser(BaseParser):
    def parse(self, source: str | bytes) -> ParseResult:
        path = source if isinstance(source, str) else _bytes_to_tmp(source, ".xlsx")
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            groups = []
            for sheet in wb.worksheets:
                groups.extend(_parse_sheet(sheet))
            confidence = _compute_confidence(groups)
            return ParseResult(groups=groups, parser_used="excel", confidence=confidence)
        except Exception as e:
            return ParseResult(groups=[], parser_used="excel", confidence=0.0,
                               warnings=[str(e)])


def _parse_sheet(sheet) -> list[dict]:
    rows = _rows_with_merged(sheet)
    if not rows:
        return []

    result = _try_mpgu_format(rows, sheet.title)
    if result:
        return result

    # Fallback: traditional format with day columns
    header_row_idx = _find_day_column_header(rows)
    if header_row_idx is None:
        return _parse_flat(rows)
    return _parse_with_day_columns(rows, header_row_idx)


# ── MPGU columnar format ──────────────────────────────────────────────────────

def _try_mpgu_format(rows: list[list[str]], sheet_title: str = "") -> list[dict]:
    """Detect and parse MPGU columnar schedule (day+time as rows, groups as columns)."""
    header = _find_mpgu_header(rows)
    if header is None:
        return []

    header_idx, day_col, time_col, data_col = header
    header_row = rows[header_idx]

    group_cols: dict[int, str] = {}
    for col_i in range(data_col, len(header_row)):
        name = header_row[col_i].strip()
        if name:
            group_cols[col_i] = name
    if not group_cols:
        return []

    # Если в строке заголовка не коды групп, а описания курсов
    # («1 курс (1 группа)»), а строкой ниже стоят реальные коды (ВVИ34-ИСТ2501) —
    # берём имена из строки с кодами и сдвигаем header_idx на неё.
    # допускаем латинские гомоглифы (V/O/I и т.п.), встречающиеся в кодах МПГУ
    _CODE = re.compile(r"[А-ЯA-Z]{2,3}\d{2}[-\s]?[А-ЯA-Z]{2,4}\s?\d{4}")
    if not any(_CODE.search(n) for n in group_cols.values()):
        for look in range(header_idx + 1, min(header_idx + 3, len(rows))):
            cand = rows[look]
            codes = {ci: cand[ci].strip() for ci in group_cols
                     if ci < len(cand) and _CODE.search(cand[ci] or "")}
            if len(codes) >= max(1, len(group_cols) // 2):
                for ci, code in codes.items():
                    group_cols[ci] = re.sub(r"\s+", "", code)
                header_idx = look
                header_row = rows[header_idx]
                break

    # Detect ZFO (date strings in day_col) vs full-time (day names)
    is_fulltime = True
    for row in rows[header_idx + 1: header_idx + 6]:
        if len(row) > day_col and row[day_col]:
            if normalize_day(row[day_col].rstrip()):
                is_fulltime = True
                break
            if re.search(r"\d{1,2}\.\d{2}\.\d{4}", row[day_col]):
                is_fulltime = False
                break

    # Detect multi-line cells (history) vs multi-row data (sport/ZFO)
    has_multiline = any(
        "\n" in (row[col_i] or "")
        for row in rows[header_idx + 1: header_idx + 20]
        for col_i in group_cols
        if col_i < len(row)
    )

    title_lower = sheet_title.lower()
    degree = "master" if any(k in title_lower for k in ("(м)", "магистр")) else "bachelor"
    form = "full_time" if is_fulltime else "correspondence"

    if has_multiline and is_fulltime:
        return _parse_mpgu_fulltime(rows, header_idx, day_col, time_col, group_cols, form, degree)
    else:
        return _parse_mpgu_multirow(rows, header_idx, day_col, time_col, group_cols, form, degree)


def _find_mpgu_header(rows: list[list[str]]) -> tuple | None:
    """Find header row; returns (idx, day_col, time_col, data_col) or None."""
    for i, row in enumerate(rows[:25]):
        a = row[0].lower() if len(row) > 0 else ""
        b = row[1].lower() if len(row) > 1 else ""
        c = row[2].lower() if len(row) > 2 else ""

        # Standard: день/группы in col B (idx 1), groups from col D (idx 3)
        if ("день" in b or "группы" in b) and any(row[j] for j in range(3, min(len(row), 15))):
            return i, 1, 2, 3

        # Shifted: день/группы in col C (idx 2), groups from col E (idx 4)
        if ("день" in c or "группы" in c) and any(row[j] for j in range(4, min(len(row), 15))):
            return i, 2, 3, 4

        # ОЗФО/ЗФО: «День недели» in col A (idx 0), «Группа / Время» in col B,
        # group codes from col C (idx 2)
        if ("день" in a) and ("группа" in b or "время" in b) \
                and any(row[j] for j in range(2, min(len(row), 15))):
            return i, 0, 1, 2

    return None


def _parse_mpgu_fulltime(
    rows, header_idx, day_col, time_col, group_cols, form, degree
) -> list[dict]:
    """Full-time format: multi-line cells, 4-row blocks per time slot."""
    groups_sched = {name: make_schedule_skeleton() for name in group_cols.values()}
    blocks: list[tuple] = []
    current_day = current_time = None
    block_rows: list[list[str]] = []

    for row in rows[header_idx + 1:]:
        if len(row) > day_col and row[day_col]:
            d = normalize_day(row[day_col].rstrip())
            if d:
                current_day = d
        if len(row) > time_col and row[time_col]:
            t = normalize_time(row[time_col])
            if t:
                if current_time is not None and current_day is not None:
                    blocks.append((current_day, current_time, block_rows))
                current_time = t
                block_rows = [row]
                continue
        if current_time is not None:
            block_rows.append(row)

    if current_time and current_day:
        blocks.append((current_day, current_time, block_rows))

    for day, times, brows in blocks:
        row0 = brows[0] if brows else []
        row2 = brows[2] if len(brows) > 2 else []

        for col_i, name in group_cols.items():
            c0 = row0[col_i] if col_i < len(row0) else ""
            c2 = row2[col_i] if col_i < len(row2) else ""

            if c0 and c2 and c0 != c2:
                lo = _parse_lesson_cell(c0, times[0], times[1])
                le = _parse_lesson_cell(c2, times[0], times[1])
                if lo:
                    groups_sched[name]["odd_week"][day].append(lo)
                if le:
                    groups_sched[name]["even_week"][day].append(le)
            elif c0:
                lesson = _parse_lesson_cell(c0, times[0], times[1])
                if lesson:
                    groups_sched[name]["odd_week"][day].append(lesson)
                    groups_sched[name]["even_week"][day].append({**lesson})
            elif c2:
                lesson = _parse_lesson_cell(c2, times[0], times[1])
                if lesson:
                    groups_sched[name]["even_week"][day].append(lesson)

    return _build_result(groups_sched, form, degree)


def _parse_mpgu_multirow(
    rows, header_idx, day_col, time_col, group_cols, form, degree
) -> list[dict]:
    """Multi-row format: lesson data spread across rows (sport + ZFO)."""
    groups_sched = {name: make_schedule_skeleton() for name in group_cols.values()}
    current_day: str | None = None
    current_time: tuple | None = None
    pending: dict[int, list[str]] = {}

    def flush():
        if not current_day or not current_time:
            return
        for col_i, lines in pending.items():
            name = group_cols.get(col_i)
            if not name or not lines:
                continue
            lesson = _parse_multirow_lines(lines, current_time[0], current_time[1])
            if lesson:
                groups_sched[name]["odd_week"][current_day].append(lesson)
                groups_sched[name]["even_week"][current_day].append({**lesson})

    for row in rows[header_idx + 1:]:
        if len(row) > day_col and row[day_col]:
            d = _day_from_date_str(row[day_col]) or normalize_day(row[day_col].rstrip())
            if d:
                current_day = d
        if len(row) > time_col and row[time_col]:
            t = normalize_time(row[time_col])
            if t:
                flush()
                pending.clear()
                current_time = t
                for col_i in group_cols:
                    if col_i < len(row) and row[col_i]:
                        pending.setdefault(col_i, []).append(row[col_i])
                continue
        if current_time is not None:
            for col_i in group_cols:
                if col_i < len(row) and row[col_i]:
                    pending.setdefault(col_i, []).append(row[col_i])

    flush()
    return _build_result(groups_sched, form, degree)


def _parse_lesson_cell(cell: str, t_start: str, t_end: str) -> dict | None:
    """Parse multi-line cell (full-time): subject\\nteacher\\nroom."""
    if not cell or cell.strip() in {"-", "–", "—", ".", ""}:
        return None
    lines = [l.strip() for l in cell.split("\n") if l.strip()]
    if not lines:
        return None

    subject_line = lines[0]
    lesson_type = normalize_lesson_type(subject_line)
    subject = re.sub(r"\s*\([ПЛ][ЗКБР]\)", "", subject_line, flags=re.I).strip(" ,.")
    subject, subgroup = extract_subgroup(subject)

    teacher = room = None
    for line in lines[1:]:
        room_m = re.search(r"\(ауд\.?\s*([^)]+)\)", line, re.I)
        if room_m and room is None:
            room = room_m.group(1).strip()
        if teacher is None and re.search(r"\b(проф|доц|асс|ст\. пр|преп)\b", line, re.I):
            teacher = re.sub(r"\(ауд\.?[^)]*\)", "", line).strip().rstrip(",. ")

    return lesson_obj(None, t_start, t_end, subject, lesson_type, teacher, room, subgroup) if subject else None


def _parse_multirow_lines(lines: list[str], t_start: str, t_end: str) -> dict | None:
    """Combine multi-row lesson data (sport/ZFO): subject → teacher → room across rows."""
    subject_parts: list[str] = []
    teacher = room = None

    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Dates-only lines: treat as notes, skip
        if re.match(r"^\d{1,2}\.\d{2}[.,]", line):
            continue
        if re.match(r"ауд\.?\s*[\d\w]", line, re.I) or re.match(r"с/з", line, re.I):
            if room is None:
                room = re.sub(r"^ауд\.?\s*", "", line, flags=re.I).strip()
            continue
        if re.search(r"\b(проф|доц|асс|ст\. преп|ст\.преп|преп)\b", line, re.I):
            if teacher is None:
                teacher = re.sub(r"\(ауд\.?[^)]*\)", "", line).strip().rstrip(",. ")
            room_m = re.search(r"\(ауд\.?\s*([^)]+)\)", line, re.I)
            if room_m and room is None:
                room = room_m.group(1).strip()
            continue
        subject_parts.append(line)

    subject = " ".join(dict.fromkeys(subject_parts)).strip()  # deduplicate consecutive repeats
    if not subject:
        return None

    lesson_type = normalize_lesson_type(subject)
    subject = re.sub(r"\s*\([ПЛ][ЗКБР]\)", "", subject, flags=re.I).strip()
    subject, subgroup = extract_subgroup(subject)
    return lesson_obj(None, t_start, t_end, subject, lesson_type, teacher, room, subgroup)


def _day_from_date_str(s: str) -> str | None:
    """Extract day from '14.02.2026 (суббота)' or pure '14.02.2026' → 'saturday'."""
    # Явное название дня в скобках
    m = re.search(r"\(([а-яёА-ЯЁ]+)\)", s)
    if m:
        day = normalize_day(m.group(1))
        if day:
            return day
    # Вычисляем по дате (ЗФО: даты без подписи дня)
    dm = re.search(r"\b(\d{1,2}\.\d{2}\.(?:\d{4}|\d{2}))\b", s)
    if dm:
        return date_str_to_weekday(dm.group(1))
    dm = re.search(r"\b(\d{1,2}\.\d{2})\b", s)
    if dm:
        return date_str_to_weekday(dm.group(1))
    return None


def _build_result(groups_sched, form, degree) -> list[dict]:
    result = []
    for name, sched in groups_sched.items():
        total = sum(len(v) for v in sched["odd_week"].values())
        if total > 0:
            result.append({
                "name": name,
                "year": None,
                "form": form,
                "degree": degree,
                "schedule": sched,
            })
    return result


# ── traditional formats (fallback) ───────────────────────────────────────────

def _find_day_column_header(rows: list[list[str]]) -> int | None:
    day_kw = {"понедельник", "вторник", "среда", "четверг", "пятница", "суббота",
               "пн", "вт", "ср", "чт", "пт", "сб"}
    for i, row in enumerate(rows[:10]):
        if sum(1 for c in row if c.lower() in day_kw) >= 3:
            return i
    return None


def _parse_with_day_columns(rows: list[list[str]], header_idx: int) -> list[dict]:
    header = rows[header_idx]
    day_cols: dict[int, str] = {}
    for i, h in enumerate(header):
        day = normalize_day(h)
        if day:
            day_cols[i] = day

    schedule = make_schedule_skeleton()
    current_time = None

    for row in rows[header_idx + 1:]:
        time_raw = _first_nonempty(row[:2])
        if time_raw:
            parsed = normalize_time(time_raw)
            if parsed:
                current_time = parsed
        if current_time is None:
            continue
        for col_i, day in day_cols.items():
            cell = row[col_i] if col_i < len(row) else ""
            if not cell:
                continue
            week_type, les = _parse_simple_cell(cell, current_time[0], current_time[1])
            if les:
                if week_type in ("odd", "both"):
                    schedule["odd_week"][day].append(les)
                if week_type in ("even", "both"):
                    schedule["even_week"][day].append({**les})

    if not any(schedule["odd_week"][d] for d in schedule["odd_week"]):
        return []
    return [{"name": "группа", "year": None, "form": "full_time",
             "degree": "bachelor", "schedule": schedule}]


def _parse_flat(rows: list[list[str]]) -> list[dict]:
    if not rows:
        return []
    header = rows[0]
    col = {
        "day": _find_col(header, ["день", "дн"]),
        "time": _find_col(header, ["время", "пара"]),
        "subject": _find_col(header, ["предмет", "дисциплина"]),
        "type": _find_col(header, ["вид", "тип"]),
        "teacher": _find_col(header, ["преподаватель", "педагог"]),
        "room": _find_col(header, ["аудитор", "кабинет"]),
    }
    if col["day"] is None or col["time"] is None:
        return []

    schedule = make_schedule_skeleton()
    for row in rows[1:]:
        day = normalize_day(_get(row, col["day"]) or "")
        if not day:
            continue
        times = normalize_time(_get(row, col["time"]) or "")
        if not times:
            continue
        subject = _get(row, col["subject"]) or ""
        lesson_type = normalize_lesson_type(_get(row, col["type"]) or subject)
        teacher = _get(row, col["teacher"])
        room = _get(row, col["room"])
        les = lesson_obj(None, times[0], times[1], subject, lesson_type, teacher, room)
        schedule["odd_week"][day].append(les)
        schedule["even_week"][day].append({**les})

    return [{"name": "группа", "year": None, "form": "full_time",
             "degree": "bachelor", "schedule": schedule}]


def _parse_simple_cell(cell: str, time_start: str, time_end: str):
    if not cell or cell in {"-", "–", "—"}:
        return "both", None
    week_m = re.search(r"\b(числ[а-я]*|знам[а-я]*|н/|з/|н\b|з\b)\b", cell, re.I)
    from scraper.normalizer.schedule_normalizer import normalize_week_type
    week_type = normalize_week_type(week_m.group(1)) if week_m else "both"
    if week_m:
        cell = cell[:week_m.start()] + cell[week_m.end():]
    lesson_type = normalize_lesson_type(cell)
    room_m = re.search(r"(?:ауд\.?\s*)?([\wА-Яа-я]-?\d{2,4})", cell, re.I)
    room = room_m.group(1) if room_m else None
    clean = re.sub(r"\bлек\b\.?|\bпрактика\b|\bпр\b\.?|\bлаб\b\.?|\bсеминар\b", "", cell, flags=re.I).strip()
    return week_type, lesson_obj(None, time_start, time_end, clean, lesson_type, None, room)


# ── helpers ───────────────────────────────────────────────────────────────────

def _cell_str(v) -> str:
    return str(v).strip() if v is not None else ""


def _rows_with_merged(sheet) -> list[list[str]]:
    """Читает лист, «разливая» объединённые ячейки.

    В .xlsx значение merged-диапазона хранится только в левой-верхней ячейке,
    остальные пустые. У МПГУ заголовки групп, дни недели и слоты времени почти
    всегда объединены по нескольким строкам/столбцам, поэтому без разливки
    парсер теряет привязку занятий к группе/дню. Копируем значение во все
    ячейки каждого merged-диапазона.
    """
    grid = [[_cell_str(c) for c in row] for row in sheet.iter_rows(values_only=True)]
    if not grid:
        return grid
    width = max((len(r) for r in grid), default=0)
    for r in grid:
        if len(r) < width:
            r.extend([""] * (width - len(r)))
    for rng in list(sheet.merged_cells.ranges):
        r0, r1 = rng.min_row - 1, rng.max_row - 1
        c0, c1 = rng.min_col - 1, rng.max_col - 1
        if r0 >= len(grid) or c0 >= width:
            continue
        val = grid[r0][c0]
        if not val:
            continue
        for ri in range(r0, min(r1, len(grid) - 1) + 1):
            for ci in range(c0, min(c1, width - 1) + 1):
                if not grid[ri][ci]:
                    grid[ri][ci] = val
    return grid


def _first_nonempty(cells: list[str]) -> str:
    return next((c for c in cells if c), "")


def _find_col(header: list[str], keywords: list[str]) -> int | None:
    for i, h in enumerate(header):
        if any(k in h.lower() for k in keywords):
            return i
    return None


def _get(row: list[str], col: int | None) -> str | None:
    if col is None or col >= len(row):
        return None
    return row[col] or None


def _compute_confidence(groups: list[dict]) -> float:
    if not groups:
        return 0.0
    total = sum(len(v) for g in groups for v in g["schedule"]["odd_week"].values())
    return min(1.0, round(total / 20, 2)) if total else 0.1


def _bytes_to_tmp(data: bytes, ext: str) -> str:
    import tempfile, os
    fd, path = tempfile.mkstemp(suffix=ext)
    os.write(fd, data)
    os.close(fd)
    return path
